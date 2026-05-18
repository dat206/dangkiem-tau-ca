"""
Excel Styles and Formatting Utilities

This module provides reusable styles and helper functions for consistent
Excel file formatting using openpyxl.
"""

from typing import List, Optional
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    NamedStyle,
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook


# ==================== Border Definition ====================
def get_thin_border() -> Border:
    """
    Create a thin black border for use in cell styles.
    
    Returns:
        Border: Border object with thin black style on all sides.
    """
    thin_side = Side(style="thin", color="000000")
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )


# ==================== Font Definition ====================
def get_default_font() -> Font:
    """
    Get the default font for the workbook.
    
    Returns:
        Font: Arial, size 11 (standard office font).
    """
    return Font(name="Arial", size=11)


# ==================== Named Styles ====================
def create_header_style() -> NamedStyle:
    """
    Create header style for column headers.
    
    Features:
    - Bold font
    - Center aligned (horizontal and vertical)
    - Text wrapping enabled
    - Gray background (D9D9D9)
    - Full border
    - Black text
    
    Returns:
        NamedStyle: Header style ready to be registered.
    """
    style = NamedStyle(name="header_style")
    style.font = Font(name="Arial", size=11, bold=True, color="000000")
    style.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    style.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    style.border = get_thin_border()
    return style


def create_data_text_style() -> NamedStyle:
    """
    Create style for text data cells.
    
    Features:
    - Full border
    - Left aligned (horizontal)
    - Center aligned (vertical)
    - Text wrapping enabled
    - No background color
    
    Returns:
        NamedStyle: Data text style ready to be registered.
    """
    style = NamedStyle(name="data_text_style")
    style.font = get_default_font()
    style.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    style.border = get_thin_border()
    return style


def create_data_number_style() -> NamedStyle:
    """
    Create style for numeric data cells.
    
    Features:
    - Full border
    - Center aligned (horizontal and vertical)
    - No background color
    - No text wrapping (numbers don't usually wrap)
    
    Returns:
        NamedStyle: Data number style ready to be registered.
    """
    style = NamedStyle(name="data_number_style")
    style.font = get_default_font()
    style.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    style.border = get_thin_border()
    return style


def create_total_style() -> NamedStyle:
    """
    Create style for total/summary rows.
    
    Features:
    - Bold font
    - Light yellow background (FFF2CC)
    - Center aligned (vertical)
    - Full border
    
    Returns:
        NamedStyle: Total style ready to be registered.
    """
    style = NamedStyle(name="total_style")
    style.font = Font(name="Arial", size=11, bold=True)
    style.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    style.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    style.border = get_thin_border()
    return style


# ==================== Style Registration ====================
def register_styles(workbook: Workbook) -> None:
    """
    Register all named styles with the workbook.
    
    This function should be called once after creating a workbook
    to make all custom styles available for use.
    
    Args:
        workbook (Workbook): The workbook to register styles with.
    
    Example:
        >>> from openpyxl import Workbook
        >>> wb = Workbook()
        >>> register_styles(wb)
        >>> ws = wb.active
        >>> ws["A1"].style = "header_style"
    """
    workbook.add_named_style(create_header_style())
    workbook.add_named_style(create_data_text_style())
    workbook.add_named_style(create_data_number_style())
    workbook.add_named_style(create_total_style())


def set_default_font(workbook: Workbook) -> None:
    """
    Set the default font for the entire workbook.
    
    This ensures all cells without explicit styling use the default font.
    
    Args:
        workbook (Workbook): The workbook to apply default font to.
    """
    workbook.font = get_default_font()


# ==================== Helper Functions ====================
def apply_auto_width(
    ws: Worksheet,
    skip_columns: Optional[List[str]] = None,
    padding: float = 2.0,
    max_width: float = 50.0,
) -> None:
    """
    Automatically adjust column widths based on content.
    
    This function calculates the optimal column width based on the longest
    content in each column, adds padding for readability, and respects
    a maximum width limit.
    
    Args:
        ws (Worksheet): The worksheet to adjust column widths for.
        skip_columns (Optional[List[str]], optional): List of column letters
            to skip (e.g., ["A", "B"]). Defaults to None.
        padding (float, optional): Extra width padding to add to each column.
            Defaults to 2.0.
        max_width (float, optional): Maximum column width to prevent columns
            from becoming too wide. Defaults to 50.0.
    
    Example:
        >>> apply_auto_width(ws, skip_columns=["A"], padding=2.5, max_width=40)
    """
    if skip_columns is None:
        skip_columns = []
    
    for column in ws.columns:
        if column[0].column_letter in skip_columns:
            continue
        
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                # Handle None values safely
                if cell.value is None:
                    cell_length = 0
                else:
                    cell_length = len(str(cell.value))
                
                if cell_length > max_length:
                    max_length = cell_length
            except (TypeError, ValueError):
                # In case of any conversion error, skip this cell
                pass
        
        # Calculate final width with padding
        adjusted_width = max_length + padding
        
        # Apply maximum width limit
        final_width = min(adjusted_width, max_width)
        
        # Set minimum width for readability
        final_width = max(final_width, 8.0)
        
        ws.column_dimensions[column_letter].width = final_width


def apply_full_border(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """
    Apply borders to an entire range of cells, including empty cells.
    
    This function ensures that all cells in the specified range have borders,
    creating a complete frame for the data table.
    
    Args:
        ws (Worksheet): The worksheet to apply borders to.
        start_row (int): Starting row number (1-indexed).
        end_row (int): Ending row number (1-indexed, inclusive).
        start_col (int): Starting column number (1-indexed).
        end_col (int): Ending column number (1-indexed, inclusive).
    
    Example:
        >>> # Apply border to A1:D10
        >>> apply_full_border(ws, start_row=1, end_row=10, start_col=1, end_col=4)
    """
    thin_border = get_thin_border()
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Ensure cells have at least default font if not already styled
            if cell.font is None or cell.font == Font():
                cell.font = get_default_font()


def apply_cell_range_style(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    style: str,
) -> None:
    """
    Apply a named style to an entire range of cells.
    
    This is a convenience function for applying a style to multiple cells at once.
    
    Args:
        ws (Worksheet): The worksheet containing the cells.
        start_row (int): Starting row number (1-indexed).
        end_row (int): Ending row number (1-indexed, inclusive).
        start_col (int): Starting column number (1-indexed).
        end_col (int): Ending column number (1-indexed, inclusive).
        style (str): The name of the style to apply (e.g., "header_style").
    
    Example:
        >>> # Apply header style to A1:D1
        >>> apply_cell_range_style(ws, 1, 1, 1, 4, "header_style")
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.style = style


# ==================== Utility Functions ====================
def column_letter_to_index(column_letter: str) -> int:
    """
    Convert column letter to 1-indexed column number.
    
    Args:
        column_letter (str): Column letter (e.g., "A", "Z", "AA").
    
    Returns:
        int: 1-indexed column number.
    
    Example:
        >>> column_letter_to_index("A")
        1
        >>> column_letter_to_index("Z")
        26
        >>> column_letter_to_index("AA")
        27
    """
    result = 0
    for char in column_letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def index_to_column_letter(column_index: int) -> str:
    """
    Convert 1-indexed column number to column letter.
    
    Args:
        column_index (int): 1-indexed column number.
    
    Returns:
        str: Column letter (e.g., "A", "Z", "AA").
    
    Example:
        >>> index_to_column_letter(1)
        "A"
        >>> index_to_column_letter(26)
        "Z"
        >>> index_to_column_letter(27)
        "AA"
    """
    result = ""
    while column_index > 0:
        column_index -= 1
        result = chr(ord("A") + column_index % 26) + result
        column_index //= 26
    return result
