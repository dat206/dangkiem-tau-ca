"""Excel Generator Service

This module handles Excel file generation for vessel reports.
Uses openpyxl for Excel creation and returns BytesIO for streaming.
"""

from io import BytesIO
from typing import List, Dict, Any
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.utils.excel_styles import (
    register_styles,
    set_default_font,
    apply_full_border,
    apply_auto_width,
)
from app.services.data_processor import get_province_name, count_for_cell, PROVINCE_MAP

# Define the standard province order for columns
PROVINCE_CODES_ORDER = ["QN", "HP", "TB", "NĐ", "NB", "TH", "NA", "HT", "QB", "QT", "QNg"]


def _to_date_object(val) -> Any:
    """Helper to convert string or date/datetime values into a datetime.date object for openpyxl"""
    if not val:
        return None
    if isinstance(val, date):
        if isinstance(val, datetime):
            return val.date()
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return val


def get_inspection_type_label(inspection_type: str) -> str:
    """Map DB inspection type to Excel column label"""
    mapping = {
        'hang_nam': 'HN',
        'tren_da': 'TĐ',
        'dinh_ky': 'ĐK',
        'giam_sat': 'GS',
        'cai_hoan': 'BT',
        'bat_thuong': 'BT',
        'lan_dau': 'ĐM',
        'dong_moi': 'ĐM'
    }
    return mapping.get(inspection_type, "")

def get_inspection_type_number(inspection_type: str) -> str:
    """Map DB inspection type to number for Bảng Kê column 'Hình thức kiểm tra'"""
    mapping = {
        'dinh_ky': '1',
        'hang_nam': '2',
        'tren_da': '3',
        'cai_hoan': '4',
        'bat_thuong': '4'
    }
    # For others like 'dong_moi', 'giam_sat', we can return empty or '1' if we want.
    return mapping.get(inspection_type, "")

def get_length_group_label(length_group: str) -> str:
    mapping = {
        'duoi_12': '<12',
        'L12_15': 'L 12-15',
        'L15_20': 'L 15-20',
        'L20_24': 'L 20-24',
        'L24_30': 'L 24-30',
        'L30_plus': 'L ≥30'
    }
    return mapping.get(length_group, "")


def generate_vessel_excel(data: List[Any]) -> BytesIO:
    """
    Generate "Bảng kê tổng hợp" (TỔNG HỢP GHI SỔ THỦ TỤC)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng hợp ghi sổ"

    register_styles(wb)
    set_default_font(wb)
    ws.sheet_view.showGridLines = False

    inspection_codes = ["HN", "TĐ", "ĐK", "GS", "BT"]
    length_group_labels = ["L 12-15", "L 15-20", "L 20-24", "L 24-30", "L ≥30"]

    headers = [
        "STT", "Ngày, tháng", "Họ và tên", "Địa chỉ", "Số đăng ký", 
        "Máy chính (KW)", "Chiều dài", "Hình thức kiểm tra", "Hạn Đk", "NGHỀ"
    ]
    headers += inspection_codes
    headers += length_group_labels
    headers += PROVINCE_CODES_ORDER

    header_row = 1
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header_text
        cell.style = "header_style"

    ws.row_dimensions[header_row].height = 28

    for row_idx, vessel in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).style = "data_number_style"

        # Date formatting for Issued Date
        cell_date = ws.cell(row=row_idx, column=2, value=_to_date_object(vessel.issued_date))
        cell_date.style = "data_text_style"
        cell_date.number_format = "dd-mm-yyyy"

        ws.cell(row=row_idx, column=3, value=vessel.owner_name or "").style = "data_text_style"
        ws.cell(row=row_idx, column=4, value=vessel.address_short or vessel.address or "").style = "data_text_style"
        
        # Registration number
        if vessel.registration_no:
            reg_no = vessel.registration_no
        else:
            prov = vessel.province_code or "UNK"
            reg_no = f"{prov}-......-TS"
        ws.cell(row=row_idx, column=5, value=reg_no).style = "data_text_style"

        # Power
        power_cell = ws.cell(row=row_idx, column=6)
        power_value = vessel.power_kw
        if power_value:
            power_cell.value = float(power_value)
            power_cell.number_format = "#,##0"
        power_cell.style = "data_number_style"

        # Length
        lmax_cell = ws.cell(row=row_idx, column=7)
        lmax_value = vessel.lmax
        if lmax_value:
            lmax_cell.value = float(lmax_value)
            lmax_cell.number_format = "#,##0.00"
        lmax_cell.style = "data_number_style"

        # Inspection Type Number (1,2,3,4)
        ws.cell(row=row_idx, column=8, value=get_inspection_type_number(vessel.inspection_type)).style = "data_text_style"
        
        # Valid Until Date
        cell_valid = ws.cell(row=row_idx, column=9, value=_to_date_object(vessel.valid_until))
        cell_valid.style = "data_text_style"
        cell_valid.number_format = "dd-mm-yyyy"
        
        ws.cell(row=row_idx, column=10, value=vessel.fishing_gear or "").style = "data_text_style"

        # Xs for inspection codes
        v_ins_lbl = get_inspection_type_label(vessel.inspection_type)
        for idx, code in enumerate(inspection_codes, start=11):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = "X" if code == v_ins_lbl else ""
            cell.style = "data_text_style"

        # Xs for length groups
        start_length_col = 11 + len(inspection_codes)
        v_lg_lbl = get_length_group_label(vessel.length_group)
        for idx, label in enumerate(length_group_labels, start=start_length_col):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = "X" if label == v_lg_lbl else ""
            cell.style = "data_text_style"

        # Xs for provinces
        start_province_col = start_length_col + len(length_group_labels)
        v_prov = (vessel.province_code or "").upper()
        # Handle QNg special match
        if v_prov == "QNG": v_prov = "QNg"
        for idx, prov in enumerate(PROVINCE_CODES_ORDER, start=start_province_col):
            cell = ws.cell(row=row_idx, column=idx)
            cell.value = "X" if prov == v_prov else ""
            cell.style = "data_text_style"

    total_row = len(data) + 2
    total_label = ws.cell(row=total_row, column=1)
    total_label.value = "Tổng số tàu"
    total_label.style = "total_style"
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)

    total_count = ws.cell(row=total_row, column=5)
    total_count.value = f"=COUNTA(E2:E{total_row - 1})"
    total_count.number_format = "#,##0"
    total_count.style = "total_style"

    for col_idx in range(6, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.style = "total_style"
        # Count Xs
        if col_idx >= 11:
            col_letter = get_column_letter(col_idx)
            cell.value = f'=COUNTIF({col_letter}2:{col_letter}{total_row - 1}, "X")'
            cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 20

    apply_full_border(ws, start_row=header_row, end_row=total_row, start_col=1, end_col=len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_quarterly_summary_excel(vessels: List[Any], quarter: int, year: int) -> BytesIO:
    """
    Generate "Báo cáo Quý theo tỉnh" - theo chuẩn file công ty.
    Cấu trúc: phần đầu hành chính (rows 1-4), header bảng 3 dòng + hàng số TT (rows 5-8),
    dữ liệu bắt đầu row 9, nhóm cột theo hình thức kiểm tra: Đóng mới, Hàng năm, Trên đà, Định kỳ, Cải hoán.
    """
    wb = Workbook()
    register_styles(wb)
    set_default_font(wb)

    provinces_in_data = sorted(list(set(v.province_code for v in vessels if v.province_code)))
    if not provinces_in_data:
        ws = wb.active
        ws.title = "No Data"
        ws.cell(row=1, column=1, value="Không có dữ liệu trong khoảng thời gian này.")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def fill_sheet(ws, is_summary: bool, target_provinces: List[str], appendix_index: int):
        ws.sheet_view.showGridLines = False

        # ============================================================
        # STYLE DEFINITIONS
        # ============================================================
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name="Times New Roman", size=11, bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        normal_font = Font(name="Times New Roman", size=11)
        bold_font = Font(name="Times New Roman", size=11, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def write_h(r, c, val):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
            return cell

        # ============================================================
        # ROW 1: Phụ lục / công văn
        # ============================================================
        ws.merge_cells("A1:U1")
        c1 = ws["A1"]
        c1.value = (
            f"Phụ lục {appendix_index}\n"
            f"(kèm theo Công văn số      /BC/ĐKTC ngày      /{quarter * 3}/{year} "
            "của Công ty CP công nghệ cao Hoàng Bảo Minh)"
        )
        c1.font = Font(name="Times New Roman", size=11, italic=True)
        c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ============================================================
        # ROW 2: Tên công ty (A2) + CỘNG HÒA (D2) + Mẫu số (Q2)
        # ============================================================
        ws.merge_cells("A2:C2")
        c2_cty = ws["A2"]
        c2_cty.value = "CÔNG TY CP CÔNG NGHỆ CAO\nHOÀNG BẢO MINH"
        c2_cty.font = bold_font
        c2_cty.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells("D2:P2")
        c2_chxhcn = ws["D2"]
        c2_chxhcn.value = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM\nĐộc Lập - Tự do - Hạnh phúc"
        c2_chxhcn.font = bold_font
        c2_chxhcn.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells("Q2:U2")
        c2_mau = ws["Q2"]
        c2_mau.value = "Mẫu số 04.BC"
        c2_mau.font = normal_font
        c2_mau.alignment = Alignment(horizontal="right", vertical="center")

        # ============================================================
        # ROW 3: Địa danh, ngày tháng năm (căn phải bên cột J)
        # ============================================================
        ws.merge_cells("J3:U3")
        c3 = ws["J3"]
        c3.value = f"Thanh Hóa, ngày      tháng {quarter * 3} năm {year}"
        c3.font = Font(name="Times New Roman", size=11, italic=True)
        c3.alignment = Alignment(horizontal="center", vertical="center")

        # ============================================================
        # ROW 4: Tiêu đề báo cáo
        # ============================================================
        ws.merge_cells("A4:U4")
        c4 = ws["A4"]
        if is_summary:
            c4.value = f"BÁO CÁO TỔNG HỢP TÌNH HÌNH ĐĂNG KIỂM TÀU CÁ QUÝ {quarter} NĂM {year}"
        else:
            c4.value = f"BÁO CÁO TÌNH HÌNH ĐĂNG KIỂM TÀU CÁ QUÝ {quarter} NĂM {year}"
        c4.font = Font(name="Times New Roman", size=14, bold=True)
        c4.alignment = Alignment(horizontal="center", vertical="center")

        # ============================================================
        # ROWS 5-7: Header bảng 3 dòng
        # ============================================================
        ws.merge_cells("A5:A7"); write_h(5, 1, "TT")
        ws.merge_cells("B5:B7"); write_h(5, 2, "Nhóm tàu")
        ws.merge_cells("C5:C7"); write_h(5, 3, "Tổng số tàu phải đăng kiểm")
        ws.merge_cells("D5:F5"); write_h(5, 4, "Số tàu theo vật liệu vỏ")
        ws.merge_cells("G5:U5"); write_h(5, 7, "Tổng số tàu cá đã đăng kiểm\n(đến thời điểm báo cáo)")

        # Row 6 - Level 2
        ws.merge_cells("D6:D7"); write_h(6, 4, "Gỗ")
        ws.merge_cells("E6:E7"); write_h(6, 5, "Thép")
        ws.merge_cells("F6:F7"); write_h(6, 6, "FRP")
        ws.merge_cells("G6:I6"); write_h(6, 7, "Số tàu đóng mới/\nlần đầu")
        ws.merge_cells("J6:L6"); write_h(6, 10, "Hàng năm")
        ws.merge_cells("M6:O6"); write_h(6, 13, "Trên đà")
        ws.merge_cells("P6:R6"); write_h(6, 16, "Định kỳ")
        ws.merge_cells("S6:U6"); write_h(6, 19, "Cải hoán")

        # Row 7 - Level 3 (Gỗ/Thép/FRP for each inspection type)
        for base_col in [7, 10, 13, 16, 19]:
            write_h(7, base_col,     "Gỗ")
            write_h(7, base_col + 1, "Thép")
            write_h(7, base_col + 2, "FRP")

        # Fill empty header cells
        for r in range(5, 8):
            for c in range(1, 22):
                cell = ws.cell(row=r, column=c)
                if not cell.value:
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.fill = header_fill

        # Row 8 - Số thứ tự cột (1)-(21)
        for c in range(1, 22):
            cell = ws.cell(row=8, column=c, value=f"({c})")
            cell.font = normal_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        # ============================================================
        # DATA ROWS (rows 9-13)
        # ============================================================
        length_groups = [
            ('L12_15',  'Lmax từ 12 ÷ < 15m'),
            ('L15_20',  'Lmax từ 15 ÷ < 20m'),
            ('L20_24',  'Lmax từ 20 ÷ < 24m'),
            ('L24_30',  'Lmax từ 24 ÷ < 30m'),
            ('L30_plus','Lmax từ ≥ 30m'),
        ]

        # (inspection_category, base_column_index)
        insp_col_map = [
            ('dong_moi', 7),   # cols G H I
            ('hang_nam', 10),  # cols J K L
            ('tren_da',  13),  # cols M N O
            ('dinh_ky',  16),  # cols P Q R
            ('cai_hoan', 19),  # cols S T U
        ]
        materials_order = ['Gỗ', 'Thép', 'FRP']

        data_start_row = 9
        for idx, (lg_code, lg_name) in enumerate(length_groups, start=1):
            r = data_start_row + idx - 1

            # Col A: STT
            cell_stt = ws.cell(row=r, column=1, value=idx)
            cell_stt.font = normal_font
            cell_stt.alignment = center_align

            # Col B: Nhóm tàu
            cell_name = ws.cell(row=r, column=2, value=lg_name)
            cell_name.font = normal_font
            cell_name.alignment = left_align

            # Cols G-U: fill inspection data by type × material
            for (insp_cat, base_col) in insp_col_map:
                for mat_idx, mat in enumerate(materials_order):
                    val = sum(count_for_cell(vessels, p_code, lg_code, mat, insp_cat) for p_code in target_provinces)
                    col = base_col + mat_idx
                    cell = ws.cell(row=r, column=col, value=val if val > 0 else "")
                    cell.font = normal_font
                    cell.alignment = center_align

            # Col D: Tổng Gỗ = SUM(G, J, M, P, S) - dùng SUM để tránh #VALUE! khi ô rỗng
            g7, g10, g13, g16, g19 = (
                get_column_letter(7), get_column_letter(10),
                get_column_letter(13), get_column_letter(16), get_column_letter(19)
            )
            cell_d = ws.cell(row=r, column=4,
                value=f"=SUM({g7}{r},{g10}{r},{g13}{r},{g16}{r},{g19}{r})")
            cell_d.font = normal_font; cell_d.alignment = center_align

            # Col E: Tổng Thép = SUM(H, K, N, Q, T)
            g8, g11, g14, g17, g20 = (
                get_column_letter(8), get_column_letter(11),
                get_column_letter(14), get_column_letter(17), get_column_letter(20)
            )
            cell_e = ws.cell(row=r, column=5,
                value=f"=SUM({g8}{r},{g11}{r},{g14}{r},{g17}{r},{g20}{r})")
            cell_e.font = normal_font; cell_e.alignment = center_align

            # Col F: Tổng FRP = SUM(I, L, O, R, U)
            g9, g12, g15, g18, g21 = (
                get_column_letter(9), get_column_letter(12),
                get_column_letter(15), get_column_letter(18), get_column_letter(21)
            )
            cell_f = ws.cell(row=r, column=6,
                value=f"=SUM({g9}{r},{g12}{r},{g15}{r},{g18}{r},{g21}{r})")
            cell_f.font = normal_font; cell_f.alignment = center_align

            # Col C: Tổng tàu phải ĐK = SUM(D, E, F)
            cell_c = ws.cell(row=r, column=3, value=f"=SUM(D{r},E{r},F{r})")
            cell_c.font = normal_font; cell_c.alignment = center_align

        # ============================================================
        # TỔNG CỘNG row
        # ============================================================
        total_row = data_start_row + len(length_groups)  # row 14
        data_end_row = total_row - 1                      # row 13

        ws.cell(row=total_row, column=1, value="").font = bold_font

        cell_tc_name = ws.cell(row=total_row, column=2, value="Tổng Cộng")
        cell_tc_name.font = bold_font
        cell_tc_name.alignment = left_align

        for c in range(3, 22):
            col_letter = get_column_letter(c)
            cell_sum = ws.cell(
                row=total_row, column=c,
                value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            )
            cell_sum.font = bold_font
            cell_sum.alignment = center_align

        # ============================================================
        # NGƯỜI LẬP (after total row)
        # ============================================================
        sign_row = total_row + 2
        cell_sign_label = ws.cell(row=sign_row, column=14, value="NGƯỜI LẬP")
        cell_sign_label.font = bold_font
        cell_sign_label.alignment = center_align

        # Placeholder for name (3 rows below label)
        ws.cell(row=sign_row + 3, column=14, value="").font = normal_font

        # ============================================================
        # BORDERS
        # ============================================================
        apply_full_border(ws, start_row=5, end_row=total_row, start_col=1, end_col=21)

        # ============================================================
        # COLUMN WIDTHS
        # ============================================================
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 9
        ws.column_dimensions['E'].width = 9
        ws.column_dimensions['F'].width = 9
        for col_idx in range(7, 22):
            ws.column_dimensions[get_column_letter(col_idx)].width = 8

        # ============================================================
        # ROW HEIGHTS
        # ============================================================
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 35
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 28
        ws.row_dimensions[5].height = 45
        ws.row_dimensions[6].height = 35
        ws.row_dimensions[7].height = 20
        ws.row_dimensions[8].height = 18

    # 1. Create the TỔNG HỢP sheet first
    ws_summary = wb.create_sheet(title="TỔNG HỢP")
    fill_sheet(ws_summary, is_summary=True, target_provinces=provinces_in_data, appendix_index=2)

    # 2. Create individual sheets for each province
    for idx, prov_code in enumerate(provinces_in_data):
        prov_name = get_province_name(prov_code).upper()
        # Valid sheet name max length is 31
        ws_prov = wb.create_sheet(title=prov_name[:31])
        fill_sheet(ws_prov, is_summary=False, target_provinces=[prov_code], appendix_index=idx + 3)

    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
